


from lead.models import Lead
from proposta.models import Perdido
from usuario.models import Usuario
from django.http import HttpResponse
from openpyxl import Workbook
import os
import tempfile
from zipfile import ZipFile
from django.http import HttpResponse
from io import BytesIO
from motor_bi.processamento import ProcessamentoBi


def criar_relatorio_excel(request):
    user_id = request.GET.get('user_id')
    motor = ProcessamentoBi()
    leads = Lead.objects.filter(user=user_id)
    prospeccoes = motor.get_prospecoes(user_id)
    propostas = motor.get_propostas_por_prospeccao(user_id)
    vendas = motor.get_vendas(user_id)
    perdidos = motor.get_perdidos(user_id)
    usuario = Usuario.objects.get(id=user_id)
    usuarios_grupos = Usuario.objects.filter(cd_grupo=usuario.cd_grupo)

    if 'data_criacao' in request.GET:
        leads = leads.filter(data_cadastro=request.GET['data_criacao'])
        prospeccoes = prospeccoes.filter(data_inicio_prospeccao=request.GET['data_criacao'])
        propostas = propostas.filter(data_cadastro=request.GET['data_criacao'])
        vendas = vendas.filter(data_fechamento=request.GET['data_criacao'])
        perdidos = perdidos.filter(data_fechamento=request.GET['data_criacao'])

    if 'data_ultima_alteracao' in request.GET:
        leads = leads.filter(data_ultima_alteracao=request.GET['data_ultima_alteracao'])
        prospeccoes = prospeccoes.filter(data_contato_inicial=request.GET['data_ultima_alteracao'])
    if 'data_proxima_acao' in request.GET:
        prospeccoes = prospeccoes.filter(data_proxima_acao=request.GET['data_proxima_acao'])
    
    if 'estagio_negocio' in request.GET:
        estagio = request.GET.get('estagio_negocio')
        if estagio == 'lead':
            excel_content = gerar_excel_leads(leads)
            return baixar_arquivo(excel_content, 'Leads.xlsx')
        elif estagio == 'prospeccao':
            excel_content = gerar_excel_prospeccoes(prospeccoes)
            return baixar_arquivo(excel_content, 'Prospecções.xlsx')
        elif estagio == 'proposta':
            excel_content = gerar_excel_propostas(propostas)
            return baixar_arquivo(excel_content, 'Propostas.xlsx')
        elif estagio == 'venda':
            excel_content = gerar_excel_transacoes(vendas, 'Vendas')
            return baixar_arquivo(excel_content, 'Vendas.xlsx')
        elif estagio == 'perdido':
            excel_content = gerar_excel_transacoes(perdidos, 'Perdido')
            return baixar_arquivo(excel_content, 'Perdidos.xlsx')

    else:
        with tempfile.TemporaryDirectory() as tmpdirname:
            leads_excel_path = os.path.join(tmpdirname, 'Leads.xlsx')
            prospeccoes_excel_path = os.path.join(tmpdirname, 'Prospeccoes.xlsx')
            propostas_excel_path = os.path.join(tmpdirname, 'Propostas.xlsx')
            vendas_excel_path = os.path.join(tmpdirname, 'Vendas.xlsx')
            perdidos_excel_path = os.path.join(tmpdirname, 'Perdidos.xlsx')
            comissao_excel_path = os.path.join(tmpdirname, 'Comissao.xlsx')

            with open(leads_excel_path, 'wb') as f:
                f.write(gerar_excel_leads(leads))
            with open(prospeccoes_excel_path, 'wb') as f:
                f.write(gerar_excel_prospeccoes(prospeccoes))
            with open(propostas_excel_path, 'wb') as f:
                f.write(gerar_excel_propostas(propostas))
            with open(vendas_excel_path, 'wb') as f:
                f.write(gerar_excel_transacoes(vendas, 'Vendas'))
            with open(perdidos_excel_path, 'wb') as f:
                f.write(gerar_excel_transacoes(perdidos, 'Perdidos'))
            with open(comissao_excel_path, 'wb') as f:
                f.write(gerar_excel_grupo(usuarios_grupos))

            zip_filename = os.path.join(tmpdirname, 'relatorio-geral.zip')
            with ZipFile(zip_filename, 'w') as zipf:
                zipf.write(leads_excel_path, 'Leads.xlsx')
                zipf.write(prospeccoes_excel_path, 'Prospeccoes.xlsx')
                zipf.write(propostas_excel_path, 'Propostas.xlsx')
                zipf.write(vendas_excel_path, 'Vendas.xlsx')
                zipf.write(perdidos_excel_path, 'Perdidos.xlsx')
                if 'comissao' in request.GET:
                    zipf.write(comissao_excel_path, 'Comissao.xlsx')

            with open(zip_filename, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename="relatorio-geral.zip"'
                return response


def gerar_excel_leads(leads):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Leads'

    lead_fields = [
        'User', 'CNPJ', 'Nome da Empresa', 'Responsável', 'Telefone',
        'Email', 'Origem', 'Cargo', 'Descrição', 'Data de Cadastro', 'Data da Última Alteração'
    ]
    ws.append(lead_fields)

    for lead in leads:
        ws.append([
            lead.user.first_name,  
            lead.cnpj,
            lead.nomeEmpresa,
            lead.responsavel,
            lead.telefone,
            lead.email,
            lead.origem,
            lead.cargo,
            lead.descricao,
            lead.data_cadastro,
            lead.data_ultima_alteracao
        ])

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)  
    return excel_file.read()


def gerar_excel_prospeccoes(prospeccoes):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Prospecções'

    prospeccao_fields = [
        'Nome do Negócio', 'Segmento', 'Serviços', 'Participação Comercial', 
        'Participação Efetiva', 'Consultor', 'Comissão', 'Data Início', 
        'Data Contato Inicial', 'Data Próxima Ação', 'Preferência Contato', 
        'Horário Contato', 'Observação', 'Status', 'Responsável', 'Versão'
    ]
    ws.append(prospeccao_fields)

    # Preencher as linhas com os valores dos objetos Prospeccao
    for prospeccao in prospeccoes:
        ws.append([
            prospeccao.nome_negocio, prospeccao.segmento, prospeccao.servicos_produtos, 
            prospeccao.participacao_comercial, prospeccao.participacao_efetiva, 
            prospeccao.consultor, prospeccao.comissao, prospeccao.data_inicio_prospeccao, 
            prospeccao.data_contato_inicial, prospeccao.data_proxima_acao, 
            prospeccao.preferencia_contato, prospeccao.horario_contato, 
            prospeccao.observacao, prospeccao.status, prospeccao.responsavel, 
            prospeccao.versao
        ])

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)  
    return excel_file.read()


def gerar_excel_propostas(propostas):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Propostas'

    proposta_fields = [
        'Nome da Proposta', 'Data de Cadastro', 'Descrição', 'Consultor', 'Tipo de Projeto',
        'Influenciador Decisor', 'Perfil Comprador', 'Probabilidade de Fechamento',
        'Status da Proposta', 'Valor da Proposta', 'Material Insumo', 'Serviços',
        'Somatório', 'Tipo de Contato', 'Versão', 'Ativa'
    ]
    ws.append(proposta_fields)

    for proposta in propostas:
        ws.append([
            proposta.nome_proposta, proposta.data_cadastro, proposta.desc_proposta,
            proposta.consultor_prop.first_name if proposta.consultor_prop else '', 
            proposta.tipo_projeto, proposta.influenciador_decisor, proposta.perfil_orcamento,
            proposta.prob_fechamento, proposta.status_proposta, proposta.valor_proposta,
            proposta.material_insumo, proposta.servicos, proposta.somatorio,
            proposta.tipo_contato, proposta.versao, proposta.ativa
        ])

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)  
    return excel_file.read()


def gerar_excel_transacoes(queryset, titulo):
    wb = Workbook()
    ws = wb.active
    ws.title = titulo

    if titulo == 'Vendas':
        campos_transacao = [
            'Data Fechamento', 'Status da Proposta', 'Valor da Proposta', 'Nome da Proposta'
        ]
        ws.append(campos_transacao)

        for transacao in queryset:
            ws.append([
                transacao.data_fechamento,
                transacao.status_proposta,
                transacao.valor_proposta,
                transacao.proposta.nome_proposta if transacao.proposta else '',
            ])
    else:
        campos_transacao = [
            'Data Fechamento', 'Status da Proposta', 'Motivo', 'Nome da Proposta'
        ]
        ws.append(campos_transacao)

        for transacao in queryset:
            ws.append([
                transacao.data_fechamento,
                transacao.status_proposta,
                transacao.motivo,
                transacao.proposta.nome_proposta if transacao.proposta else '',
            ])
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file.read()

def gerar_excel_grupo(usuarios):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Grupo'

    grupo_fields = [
        'Nome', 'Email', 'Comissao'
    ]
    ws.append(grupo_fields)

    for usuario in usuarios:
        ws.append([
            usuario.first_name, usuario.email, 'Tem comissao' if usuario.comissao else 'Nao tem comissao'
        ])

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)  
    return excel_file.read()


def baixar_arquivo(excel_content, nome_arquivo):
    response = HttpResponse(excel_content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    return response